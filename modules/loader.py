"""
loader.py — Carga incremental de datos desde archivos Excel hacia SQLite
Clave de join entre ambas fuentes: RUT
"""
import os
import re
import pandas as pd
import openpyxl
from datetime import datetime
from typing import Optional
from modules.db import (
    init_db, upsert_dotacion, insert_licencias,
    upsert_gestion, log_carga, TIPOS_EXCLUIDOS
)

# ────────────────────────────────────────────────────────────────────────────
# Tipos de LM válidas (para clasificar categoría)
# ────────────────────────────────────────────────────────────────────────────
PREFIJOS_LM = ('L.M.', 'LM ', 'LICENCIA MEDICA', 'LICENCIA MÉDICA', 'PERMISO SANNA')

def _es_lm(tipo: str) -> bool:
    t = str(tipo).upper().strip()
    return any(t.startswith(p) for p in PREFIJOS_LM)

def _categoria(tipo: str) -> str:
    """
    Retorna:
      'LM'       → Licencia Médica válida (se incluye en todos los índices)
      'EXCLUIDO' → Item expresamente excluido por CGR (no se ingresa a la BD)
    La base de datos SOLO contiene Licencias Médicas.
    """
    t = str(tipo).upper().strip()
    if t in {x.upper() for x in TIPOS_EXCLUIDOS}:
        return 'EXCLUIDO'
    return 'LM'

# ────────────────────────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────────────────────────
def _clean_rut(val) -> Optional[int]:
    try:
        return int(str(val).strip().replace('.', '').replace('-', '').split()[0])
    except Exception:
        return None

def _clean_str(val) -> str:
    if val is None:
        return ''
    return str(val).strip()

def _clean_date(val) -> str:
    if val is None or str(val).strip() in ('', 'None', 'NaT', '00/00/0000'):
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    # DD/MM/YYYY
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    # YYYY-MM-DD
    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
        return s[:10]
    return s

def _clean_float(val) -> Optional[float]:
    try:
        return float(str(val).replace('$', '').replace(',', '').strip())
    except Exception:
        return None

def _clean_int(val) -> Optional[int]:
    try:
        return int(float(str(val).strip()))
    except Exception:
        return None

def _load_wb_sheet(file_path: str, sheet_name: str) -> list:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise KeyError(
            f"Hoja '{sheet_name}' no encontrada en el archivo.\n"
            f"Hojas disponibles: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    rows = [row for row in ws.iter_rows(values_only=True)
            if any(c is not None for c in row)]
    wb.close()
    return rows


def _find_sheet(wb, candidates: list, fallback_contains: list = None) -> Optional[str]:
    """
    Busca la primera hoja cuyo nombre (strip, upper) coincida con algún candidato.
    Si no encuentra, busca por contenido parcial (fallback_contains).
    Retorna None si no encuentra nada.
    """
    names_upper = {s.strip(): s.strip().upper() for s in wb.sheetnames}
    cands_upper = [c.upper() for c in candidates]

    # Coincidencia exacta (case-insensitive)
    for sheet, upper in names_upper.items():
        if upper in cands_upper:
            return sheet

    # Coincidencia parcial
    if fallback_contains:
        for sheet, upper in names_upper.items():
            if any(k.upper() in upper for k in fallback_contains):
                return sheet

    return None

# ────────────────────────────────────────────────────────────────────────────
# Carga Dotación desde archivos de dotación
# ────────────────────────────────────────────────────────────────────────────
# Nombres conocidos para la hoja principal de dotación
_DOT_SIN_DUP = ['SIN DUPLICADOS', 'DOT LIMPIA', 'DOT_LIMPIA', 'TIT-CONTRATA MARZO 2026',
                'TIT-CONTRA MARZO 2026', 'TIT-CONT', 'DOTACION DAP']
# Nombres conocidos para la hoja extendida (con datos contractuales extra)
_DOT_DUP2 = ['DUPLICADOS (2)', 'DUPLICADOS', 'DOTACION NOVIEMBRE 30-11-20 (2)',
             'TIT-CONTRATA MARZO 2026 (2)', 'TIT-CONTRA (2)', 'HSA MARZO 2026']


def load_dotacion(file_path: str) -> dict:
    """
    Carga dotación. Detecta automáticamente los nombres reales de las hojas
    (compatible con distintas versiones del archivo de dotación DAP).
    """
    init_db()
    archivo = os.path.basename(file_path)

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_sin = _find_sheet(wb, _DOT_SIN_DUP, fallback_contains=['SIN DUPLIC', 'DOT'])
    sheet_dup2 = _find_sheet(wb, _DOT_DUP2, fallback_contains=['DUPLIC'])
    wb.close()

    if not sheet_sin:
        raise ValueError(
            f"No se encontró hoja de dotación en '{archivo}'.\n"
            f"Se esperaba una de: {_DOT_SIN_DUP}"
        )

    # Hoja principal
    rows_sin = _load_wb_sheet(file_path, sheet_sin)
    header_sin = rows_sin[0]

    # Hoja extendida (opcional — datos contractuales adicionales)
    rows_dup2 = _load_wb_sheet(file_path, sheet_dup2) if sheet_dup2 else []
    header_dup2 = rows_dup2[0] if rows_dup2 else []

    # Índices hoja extendida (75 cols)
    idx2 = {str(h).strip(): i for i, h in enumerate(header_dup2) if h}

    # Construir lookup por RUT desde DUPLICADOS (2)
    lookup_dup2 = {}
    for row in rows_dup2[1:]:
        rut = _clean_rut(row[0])
        if rut:
            lookup_dup2[rut] = row

    records = []
    for row in rows_sin[1:]:
        rut = _clean_rut(row[0])
        if not rut:
            continue
        ext = lookup_dup2.get(rut, [None]*75)

        def g2(campo):
            i = idx2.get(campo)
            return ext[i] if i is not None and i < len(ext) else None

        rec = {
            'rut':              rut,
            'dv':               _clean_str(row[1]),
            'nombre':           _clean_str(row[2]),
            'planta':           _clean_str(row[3]),
            'calidad_juridica': _clean_str(row[4]),
            'genero':           _clean_str(row[5]),
            'ley':              _clean_str(row[6]),
            'horas':            _clean_int(row[7]),
            'est_codigo':       _clean_int(row[8]),
            'est_nombre':       _clean_str(row[9]),
            'fecha_nacimiento': _clean_date(row[10]),
            'cod_unidad':       _clean_int(row[11]),
            'unidad':           _clean_str(row[12]),
            'cargo':            _clean_str(row[13]),
            'funcion':          _clean_str(row[14]),
            'titulo':           _clean_str(row[15]),
            'fecha_ini_contrato': _clean_date(row[16]),
            'fecha_ter_contrato': _clean_date(row[17]),
            # Desde DUPLICADOS (2)
            'direccion':        _clean_str(g2('Dirección')),
            'ciudad':           _clean_str(g2('Ciudad')),
            'comuna':           _clean_str(g2('Comuna')),
            'fecha_ingreso_servicio': _clean_date(g2('Fecha Ingreso Servicio')),
            'isapre':           _clean_str(g2('Descripción Isapre')),
            'afp':              _clean_str(g2('Descripción')),
            'cargas_familiares': _clean_int(g2('Cargas Familiares')),
            'remuneracion':     _clean_float(g2('Remuneración')),
        }
        records.append(rec)

    ins, upd = upsert_dotacion(records)
    hojas_usadas = sheet_sin + (f' + {sheet_dup2}' if sheet_dup2 else '')
    carga_id = log_carga(
        archivo, hojas_usadas, 'dotacion',
        ins, upd, notas=f'{len(records)} registros procesados'
    )
    return {'insertados': ins, 'actualizados': upd, 'total': len(records), 'carga_id': carga_id}


# ────────────────────────────────────────────────────────────────────────────
# Carga Licencias desde BASE AUSENTISMO (hoja 31.03.2026 o similar)
# ────────────────────────────────────────────────────────────────────────────
def load_licencias(file_path: str, sheet_name: str = None) -> dict:
    """
    Carga registros de LM desde la hoja transaccional.
    Si sheet_name es None, detecta automáticamente la hoja con más filas.
    """
    init_db()
    archivo = os.path.basename(file_path)

    # Auto-detectar hoja si no se especifica
    if sheet_name is None:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        # Buscar hoja con fecha (patrón DD.MM.YYYY) o la de mayor cantidad de filas
        sheet_name = _detect_lm_sheet(wb)
        wb.close()

    rows = _load_wb_sheet(file_path, sheet_name)
    if not rows:
        return {'insertados': 0, 'omitidos': 0, 'total': 0}

    header = rows[0]
    idx = {str(h).strip().upper().replace(' ', '_'): i
           for i, h in enumerate(header) if h}

    def g(row, *names):
        for n in names:
            i = idx.get(n)
            if i is not None and i < len(row):
                return row[i]
        return None

    records = []
    fechas = []
    for row in rows[1:]:
        rut = _clean_rut(g(row, 'RUT', '_RUT'))
        if not rut:
            continue

        tipo_raw = _clean_str(g(row, 'TIPO_DE_AUSENTISMO', 'TIPO DE AUSENTISMO'))
        cat = _categoria(tipo_raw)
        if cat == 'EXCLUIDO':
            continue  # No insertar items CGR

        fi = _clean_date(g(row, 'F.INICIO', 'FECHA_INICIO', 'F._INICIO'))
        ft = _clean_date(g(row, 'F.TERMINO', 'FECHA_TERMINO', 'F._TERMINO'))

        if fi:
            fechas.append(fi)

        rec = {
            'rut':          rut,
            'dv':           _clean_str(g(row, 'DV', '_DV')),
            'nombre':       _clean_str(g(row, 'NOMBRE', '_NOMBRE')),
            'ley':          _clean_str(g(row, 'LEY', '_LEY')),
            'edad_anos':    _clean_int(g(row, 'EDAD_(AÑOS)', 'EDAD_(A\xd1OS)', 'EDAD_A\xd1OS')),
            'afp':          _clean_str(g(row, 'AFP', '_AFP')),
            'salud':        _clean_str(g(row, 'SALUD', '_SALUD')),
            'cod_unidad':   _clean_int(g(row, 'CODIGO_UNIDAD', '_CODIGO_UNIDAD')),
            'nombre_unidad': _clean_str(g(row, 'NOMBRE_UNIDAD', '_NOMBRE_UNIDAD')),
            'genero':       _clean_str(g(row, 'GENERO', '_GENERO')),
            'cargo':        _clean_str(g(row, 'CARGO', '_CARGO')),
            'calidad_juridica': _clean_str(g(row, 'CALIDAD_JURIDICA', '_CALIDAD_JURIDICA')),
            'planta':       _clean_str(g(row, 'PLANTA', '_PLANTA')),
            'num_resolucion': _clean_str(g(row, 'N°_RESOLUCION', 'N°_RESOLUCION', 'Nº_RESOLUCION')),
            'fecha_resolucion': _clean_date(g(row, 'F._RESOLUCION', 'FECHA_RESOLUCION')),
            'fecha_inicio': fi,
            'fecha_termino': ft,
            'total_dias':   _clean_int(g(row, 'TOTAL_DIAS_AUSENTISMO', 'TOTAL_DIAS')),
            'dias_periodo': _clean_int(g(row, 'AUSENTISMO_EN_EL_PERIODO', 'DIAS_PERIODO')),
            'costo':        _clean_float(g(row, 'COSTO_DE_LICENCIA', 'COSTO')),
            'tipo_lm':      tipo_raw,
            'cod_establecimiento': _clean_int(g(row, 'CODIGO_DE_ESTABLECIMIENTO')),
            'nombre_establecimiento': _clean_str(g(row, 'NOMBRE_DE_ESTABLECIMIENTO')),
            'saldo_dias_no_reemplazados': _clean_int(g(row, 'SALDO_DIAS_NO_REEMPLAZADOS')),
            'tipo_contrato': _clean_str(g(row, 'TIPO_DE_CONTRATO', '_TIPO_DE_CONTRATO')),
            'categoria':    cat,
        }
        records.append(rec)

    periodo_ini = min(fechas) if fechas else ''
    periodo_fin = max(fechas) if fechas else ''

    # Carga incremental: insertar solo registros nuevos (UNIQUE constraint)
    carga_id = log_carga(archivo, sheet_name, 'licencias', 0, 0,
                          periodo_ini, periodo_fin, 'pre-insercion')
    ins, skip = insert_licencias(records, carga_id)

    # Actualizar log con counts reales
    from modules.db import execute
    execute("UPDATE cargas_log SET registros_nuevos=?, notas=? WHERE id=?",
            (ins, f'{ins} nuevos, {skip} duplicados omitidos', carga_id))

    return {
        'insertados': ins, 'omitidos': skip,
        'total': len(records), 'carga_id': carga_id,
        'periodo': f'{periodo_ini} → {periodo_fin}'
    }


_LM_SHEET_NAMES = ['SALIDA', '31.03.2026', 'LM', 'LICENCIAS', 'BASE LM', 'AUSENTISMOS']
_LM_KEYWORDS = ['SALIDA', 'LM', 'LICENCIA', 'AUSENTISMO']


def _detect_lm_sheet(wb) -> str:
    """
    Detecta la hoja transaccional de LM.
    Orden de prioridad:
      1. Patrón de fecha DD.MM.YYYY
      2. Nombre exacto conocido (SALIDA, LM, etc.)
      3. Nombre con keyword parcial
      4. Hoja con más filas (fallback final)
    """
    fecha_pat = re.compile(r'^\d{2}\.\d{2}\.\d{4}$')
    names_upper = {s.strip(): s.strip().upper() for s in wb.sheetnames}

    # 1. Patrón de fecha
    for sheet in names_upper:
        if fecha_pat.match(sheet):
            return sheet

    # 2. Nombre exacto conocido
    for sheet, upper in names_upper.items():
        if upper in _LM_SHEET_NAMES:
            return sheet

    # 3. Keyword parcial (excluir hoja AUSENTISMO de gestión de casos)
    for sheet, upper in names_upper.items():
        if upper == 'AUSENTISMO':
            continue  # Esa es la hoja de gestión, no de LM
        if any(k in upper for k in _LM_KEYWORDS):
            return sheet

    # 4. Hoja con más filas
    best, best_count = wb.sheetnames[0], 0
    for name in wb.sheetnames:
        ws = wb[name]
        count = ws.max_row or 0
        if count > best_count:
            best, best_count = name, count
    return best


# ────────────────────────────────────────────────────────────────────────────
# Carga Gestión de Casos (hoja AUSENTISMO)
# ────────────────────────────────────────────────────────────────────────────
def load_gestion(file_path: str, sheet_name: str = 'AUSENTISMO') -> dict:
    init_db()
    archivo = os.path.basename(file_path)
    rows = _load_wb_sheet(file_path, sheet_name)
    if not rows:
        return {'insertados': 0, 'actualizados': 0, 'total': 0}

    header = [_clean_str(h) for h in rows[0]]

    def col(row, *names):
        for n in names:
            try:
                i = header.index(n)
                return row[i] if i < len(row) else None
            except ValueError:
                continue
        return None

    records = []
    for row in rows[1:]:
        rut = _clean_rut(col(row, 'RUT'))
        if not rut:
            continue
        # Las columnas de ventanas rolling (MAY23 ABR25, etc.) las ignoramos aquí
        rec = {
            'rut':              rut,
            'nombre':           _clean_str(col(row, 'NOMBRE')),
            'vigente':          _clean_str(col(row, 'VIGENTES')),
            'proceso':          _clean_str(col(row, 'RESUMEN DEL PROCESO')),
            'resumen_interno':  _clean_str(col(row, 'RESUMEN INTERNO')),
            'fecha_usf':        _clean_date(col(row, ' FECHA USF')),
            'compin':           _clean_str(col(row, 'COMPIN')),
            'casos_especiales': _clean_str(col(row, 'CASOS ESPECIALES')),
            'diferencial':      _clean_float(col(row, 'DIFERENCIAL')),
            'promedio':         _clean_float(col(row, 'PROMEDIO')),
            'aus_ajustado':     _clean_float(col(row, 'AUS. AJUSTADO')),
            'ult_ausentismo':   _clean_float(col(row, 'ÚLT. AUSENTISMO', '\xdaLT. AUSENTISMO')),
            'genero':           _clean_str(col(row, 'Género', 'G\xe9nero')),
            'fecha_nacimiento': _clean_date(col(row, 'Fecha Nacimiento')),
            'establecimiento':  _clean_str(col(row, 'ESTABLECIMIENTO')),
            'estamento':        _clean_str(col(row, 'ESTAMENTO')),
            'responsable':      _clean_str(col(row, 'RESPONSABLE')),
            'asignacion':       _clean_date(col(row, 'ASIGNACIÓN', 'ASIGNACI\xd3N')),
            'visita_dom':       _clean_str(col(row, 'VISITA DOMICILIARIA')),
            'fecha_comite':     _clean_date(col(row, 'FECHA COMITE')),
            'carta':            _clean_str(col(row, 'CARTA')),
            'n_contacto':       _clean_str(col(row, 'N° DE CONTACTO')),
            'correo':           _clean_str(col(row, 'correo')),
            'observaciones':    '',  # columna sin header fijo
        }
        records.append(rec)

    ins, upd = upsert_gestion(records)
    carga_id = log_carga(archivo, sheet_name, 'gestion', ins, upd,
                          notas=f'{len(records)} registros procesados')
    return {'insertados': ins, 'actualizados': upd, 'total': len(records), 'carga_id': carga_id}


# ────────────────────────────────────────────────────────────────────────────
# Carga inteligente: detecta qué hojas cargar según el archivo
# ────────────────────────────────────────────────────────────────────────────
def smart_load(file_path: str, progress_cb=None) -> list[dict]:
    """
    Detecta automáticamente el tipo de archivo y carga todas las hojas relevantes.
    Retorna lista de resultados por hoja.
    """
    init_db()
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets = [s.strip() for s in wb.sheetnames]
    wb.close()

    results = []

    # ¿Es archivo de dotación? (detectar por nombres conocidos de hoja)
    sheets_upper = [s.strip().upper() for s in sheets]
    es_dotacion = any(s in sheets_upper for s in [n.upper() for n in _DOT_SIN_DUP])
    if es_dotacion:
        if progress_cb:
            progress_cb(0.2, 'Cargando dotación...')
        r = load_dotacion(file_path)
        r['hoja'] = 'Dotación (hojas detectadas automáticamente)'
        results.append(r)

    # ¿Tiene hoja de gestión?
    if 'AUSENTISMO' in sheets:
        if progress_cb:
            progress_cb(0.5, 'Cargando gestión de casos...')
        r = load_gestion(file_path, 'AUSENTISMO')
        r['hoja'] = 'Gestión de Casos (AUSENTISMO)'
        results.append(r)

    # ¿Tiene hoja transaccional de LM?
    fecha_pat = re.compile(r'^\d{2}\.\d{2}\.\d{4}$')
    lm_sheet = next((s for s in sheets if fecha_pat.match(s.strip())), None)
    if not lm_sheet:
        lm_sheet = next(
            (s for s in sheets
             if s.strip().upper() in _LM_SHEET_NAMES
             and s.strip().upper() != 'AUSENTISMO'), None
        )
    if not lm_sheet:
        lm_sheet = next(
            (s for s in sheets
             if any(k in s.upper() for k in _LM_KEYWORDS)
             and s.strip().upper() != 'AUSENTISMO'), None
        )
    if lm_sheet:
        if progress_cb:
            progress_cb(0.7, f'Cargando licencias ({lm_sheet})...')
        r = load_licencias(file_path, lm_sheet)
        r['hoja'] = f'Licencias ({lm_sheet})'
        results.append(r)

    if progress_cb:
        progress_cb(1.0, 'Carga completada')
    return results
